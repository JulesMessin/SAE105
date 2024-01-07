"""import sae5_module.py"""
import openpyxl
import matplotlib.pyplot as plt
import statistics


#Je n'ai pas mis de fonction pour toute cette partie car cela provoque une erreur que je n'ai pas su résoudre à la génération du graphique (crée une sorte de griboullage sur le graphique obtenue)


#Lecture du fichier excel et enregistrement des données dans les variables value_list et date_list
workbook = openpyxl.load_workbook("../data/conductivite_amont.xlsx", data_only = True)
titres_onglets = workbook.sheetnames
onglet1 = workbook[titres_onglets[0]]
min_row_o1 = onglet1.min_row
max_row_o1 = onglet1.max_row
min_col_o1 = onglet1.min_column
max_col_o1 = onglet1.max_column
for col in onglet1.iter_cols(min_row = min_row_o1, max_row = max_row_o1, min_col = min_col_o1, max_col = max_col_o1, values_only=True):
    if col[0]=='Value':
        value_list=col
    elif col[0]=='Date/Time':
        date_list=col

#conversion de la liste value_list en list, puis on enlève la première valeur de la liste qui est "Valeur"
value_list=list(value_list)
value_list.pop(0)

#conversion de la liste date_list en list, puis on enlève la première valeur de la liste qui est "Date/Time"
date_list=list(date_list)
date_list.pop(0)

#On ferme le fichier excel
workbook.close()

#Création d'une variable date_s_list à partir de la liste date_list, permet d'obtenir le temps en s en fonction du nombre de valeur
date_s_list=[]
for i in range(len(date_list)):
    date_s_list.append(i)
date_s_list=len(date_s_list)/2

# Creation d'un figure
fig = plt.plot(value_list)
plt.xlim(0,date_s_list)
plt.xlabel("Temps en s")
plt.ylabel("Concentration de sel en cm")
plt.title("Concentration en sel au cours du temps")
plt.savefig("../html/img/graphique.png")
plt.show()

def genere_html(filename, titre_page, body_html):
    """
    Cette fonction va permettre de générer le fichier html.

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    titre_page : HTML
        Permet d'afficher le titre de la page.
    body_html : HTML
        Contient le code qui permet d'afficher la page html.

    Returns
    -------

    """
    fichier = open(filename, 'w')
    fichier.write("".join(body_html))
    fichier.close()

def genere_css(filename, titre_page, body_css):
    """
    Cette fonction va permettre de générer le fichier css.    

    Parameters
    ----------
    filename : TYPE
        DESCRIPTION.
    titre_page : TYPE

    body_css : CSS
        Contient le code css qui modifie la page html

    Returns
    -------

    """
    fichier = open(filename, 'w')
    fichier.write("".join(body_css))
    fichier.close()

def main(calcul):
    """
    Création de la structure de la page html.

    Returns
    -------

    """
    body_html = """
                <!DOCTYPE html>
                <html>
                    <head>
                        <link rel="stylesheet" href="stylesheet.css">
                    </head>
                    <body>
                        <nav>
                            <a href="https://www.univ-poitiers.fr/"><img src="img/universite-poitiers-logo.jpg" alt="logo"></a>
                        </nav>
    
                        <div class="description">
                            <p class="up">Réalisé par Jules Messin et Benjamin Vigner.</p>
                            <p>Bonjour et bienvenue sur notre projet de la SAE105 intitulé "calcul du débit d’une rivière par mesure de concentration en sel".
                            <br><br><br>L'objectif de ce projet est de déterminer le débit de la rivière en utilisant les concentrations de sel mesurées grâce à une sonde.
                            <br><br>Nous devons donc concevoir un site web (HTML/CSS) à partir d'un fichier Python contenant le calcul du débit <br><br> et un graphique représentant la concentration en sel sur une période de temps.</p>
                        </div>
                        <div class="resultat">
                            <div class="box">
                                <p>Voici la formule :</p>
                                <img src="img/formule.png">
                                <p>Le résultat du calcul est :</p>
                                <p>""",str(calcul),"""</p>
                            
                            </div>
                            <p1>Voici la méthode de calcul du débit basée sur les concentrations en sel obtenues à partir des données fournies. </p1>
                            <img class="first" src="img/graphique.png" alt="logo">
                            <p2>Voici le graphique illustrant l'évolution des concentrations en sel au fil du temps.</p2>
                        </div>
                    </body>
                </html>
           """
    genere_html("../html/index.html", "mon titre", body_html)

def stylesheet():
    """
    Création de la structure de la page css.

    Returns
    -------

    """
    body_css = """
                body {
                  background-color: #E9E9E9;
                }

                nav {
                  background-color: white;
                  width: 91%;
                  margin-left: 4%;
                  margin-right: 5%;
                  height: 75px;
                  position: fixed;
                  margin-top: 10px;
                  z-index: 5;
                  display: inline;
                  border-radius: 10px;
                  border: 1px solid black;
                }

                div.description {
                  background-size: cover;
                  background-repeat: no-repeat;
                  height: 100%;
                  width: 100%;
                  position: absolute;
                  top: 0px;
                  left: 0px;
                  right: 0px;
                  text-align: center;
                }

                div.description p {
                  margin-top: 14%;
                  color: black;
                  font-family: monospace;
                  font-weight: bolder;
                  font-size: 150%;
                }
                div.description p.up {
                  margin-top: 6%;
                  color: black;
                  font-family: monospace;
                  font-weight: bolder;
                  font-size: 100%;
                }

                div.resultat {
                  background-color: #E9E9E9;
                  height: 1000px;
                  width: 100%;
                  position: absolute;
                  top: 100%;
                  left: 0px;
                  right: 0px;
                  text-align: center;
                  justify-content: center;
                }

                nav a img{
                  position: absolute;
                  width: 100px;
                  height: 70px;
                  top: 50%;
                  left: 50%;
                  transform: translate(-50%, -50%);
                }

                div.resultat img.first {
                  position: absolute;
                  top: 60%;
                  left: 75%;
                  transform: translate(-50%, -50%);
                  width: 700px;
                  height: 500px;
                  border-radius: 30px;
                  border: 1px solid black;
                }
                div.resultat div.box {
                  background-color: white;
                  position: absolute;
                  top: 60%;
                  left: 25%;
                  transform: translate(-50%, -50%);
                  width: 700px;
                  height: 500px;
                  border-radius: 30px;
                  border: 1px solid black;
                  font-color: black;
                  font-family: monospace;
                  font-size: 20px;
                  font-weight: bolder; 
                  font-style: italic;
                }
                p1 {
                  position: absolute;  
                  width: 700px;
                  top: 27.5%;
                  left: 25%;
                  transform: translate(-50%, -50%);  
                  color: black;
                  font-family: monospace;
                  font-size: 20px;
                  font-weight: bolder;
                
                }
                p2 {
                  position: absolute;  
                  width: 700px;
                  top: 27.5%;
                  left: 75%;
                  transform: translate(-50%, -50%);
                  color: black;
                  font-family: monospace;
                  font-size: 20px;
                  font-weight: bolder;
                }
                
               
           """
    genere_css("../html/stylesheet.css", "mon titre", body_css)

#Fonction permettant de calculer le débit d'une riviére
def calcul_debit():
    Masse = 1037
    T = date_s_list
    Cm = statistics.mean(value_list)
    Ci = value_list[0]
    deltaC = Cm - Ci
    NaCi = deltaC/2
    return Masse/(NaCi*T)

if __name__ == "__main__":
    calculx=(calcul_debit())
    main(calculx)
    stylesheet()
